# ============================================================
# ui/tabs/ar_tab.py  —  미수 / 수금 관리 탭
# ============================================================

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QComboBox, QDateEdit, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QMessageBox, QTabWidget, QFrame, QFileDialog
)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QColor
from datetime import date as dt

from logic.ar_logic import ARLogic
from logic.customer_logic import CustomerLogic
from ui.dialogs.ar_dialog import ARCollectionDialog, InitialARDialog
from ui.widgets import StatCard, SectionCard, NoScrollDateEdit
from utils.format_utils import fmt_currency, fmt_number
from database.models import DISTRICTS


class ARTab(QWidget):
    """미수 / 수금 관리 탭."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.ar_logic       = ARLogic()
        self.customer_logic = CustomerLogic()
        self._col_ids       = []
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 18, 24, 18)
        root.setSpacing(14)

        # ══ 요약 카드 ════════════════════════════════════════
        cards_h = QHBoxLayout(); cards_h.setSpacing(12)
        self.sc_total    = StatCard("💸", "총 미수 잔액",  "—", "#ef4444")
        self.sc_cust_cnt = StatCard("🏢", "미수 거래처 수", "—", "#f59e0b")
        self.sc_monthly  = StatCard("📅", "이번 달 수금",  "—", "#10b981")
        for sc in (self.sc_total, self.sc_cust_cnt, self.sc_monthly):
            cards_h.addWidget(sc)
        cards_h.addStretch()
        root.addLayout(cards_h)

        # ══ 내부 탭 ══════════════════════════════════════════
        inner = QTabWidget()
        inner.addTab(self._outstanding_tab(), "📋  미수 현황")
        inner.addTab(self._collection_tab(),  "💵  수금 내역")
        root.addWidget(inner)

        self._load_summary()

    # ── 미수 현황 탭 ─────────────────────────────────────────
    def _outstanding_tab(self) -> QWidget:
        w = QWidget()
        v = QVBoxLayout(w)
        v.setContentsMargins(16, 14, 16, 14)
        v.setSpacing(12)

        # 필터 + 버튼
        fh = QHBoxLayout(); fh.setSpacing(8)
        self.combo_dist_filter = QComboBox()
        self.combo_dist_filter.setFixedWidth(130)
        self.combo_dist_filter.addItem("전체 지역", "")
        for d in DISTRICTS:
            self.combo_dist_filter.addItem(d, d)
        self.combo_dist_filter.currentIndexChanged.connect(self._load_outstanding)
        fh.addWidget(self.combo_dist_filter)

        btn_ref = QPushButton("새로고침")
        btn_ref.setObjectName("btn_ghost")
        btn_ref.setFixedWidth(100)
        btn_ref.clicked.connect(self._load_outstanding)
        fh.addWidget(btn_ref)
        fh.addStretch()

        btn_collect = QPushButton("수금 등록")
        btn_collect.setObjectName("btn_success")
        btn_collect.clicked.connect(lambda: self._add_collection(None))
        fh.addWidget(btn_collect)

        btn_collect_sel = QPushButton("선택 거래처 수금")
        btn_collect_sel.setObjectName("btn_warning")
        btn_collect_sel.clicked.connect(self._add_collection_selected)
        fh.addWidget(btn_collect_sel)

        btn_init_ar = QPushButton("⚖️ 초기(이월) 미수 조정")
        btn_init_ar.setObjectName("btn_ghost")
        btn_init_ar.setStyleSheet("QPushButton { font-weight: bold; color: #ea580c; border: 1px solid #fdba74; background: #fff7ed; } QPushButton:hover { background: #ffedd5; }")
        btn_init_ar.clicked.connect(self._adjust_initial_ar)
        fh.addWidget(btn_init_ar)
        
        btn_export = QPushButton("📥 엑셀 다운로드")
        btn_export.setObjectName("btn_success")
        btn_export.clicked.connect(self._export_ar_excel)
        fh.addWidget(btn_export)

        btn_print = QPushButton("🖨️ 인쇄")
        btn_print.setObjectName("btn_ghost")
        btn_print.setStyleSheet("QPushButton { font-weight: bold; color: #4338ca; border: 1px solid #a5b4fc; background: #eef2ff; } QPushButton:hover { background: #e0e7ff; }")
        btn_print.clicked.connect(self._print_ar_list)
        fh.addWidget(btn_print)
        
        v.addLayout(fh)

        c = SectionCard("거래처별 미수 잔액")
        self.tbl_out = self._tbl(
            ['거래처', '지역', '초기(이월) 미수', '판매 미수액', '총 수금액', '잔여 미수']
        )
        hh = self.tbl_out.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.Stretch)
        hh.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.tbl_out_cust_ids = []
        c.add_widget(self.tbl_out)
        v.addWidget(c)

        # 거래처 클릭 시 → 미수 판매 상세 표시
        self.tbl_out.itemSelectionChanged.connect(self._on_select_customer)
        # 거래처 더블클릭 시 → 날짜별 그룹 상세
        self.tbl_out.cellDoubleClicked.connect(self._on_double_click_customer)

        detail_c = SectionCard("미수 판매 상세 내역 (선택한 거래처)")
        self.tbl_credit_det = self._tbl(
            ['날짜', '종류', '규격', '수량', '금액']
        )
        detail_c.add_widget(self.tbl_credit_det)
        v.addWidget(detail_c)

        # 날짜별 그룹 상세 (tab 1 더블클릭 시)
        date_grp_c = SectionCard("📅 날짜별 미수 금액 (더블클릭 시 표시)")
        self.tbl_date_group = self._tbl(
            ['납품일자', '품목 수', '총 수량', '미수 금액']
        )
        date_grp_c.add_widget(self.tbl_date_group)
        v.addWidget(date_grp_c)

        self._load_outstanding()
        return w

    def _load_outstanding(self):
        dist = self.combo_dist_filter.currentData()
        rows = self.ar_logic.get_outstanding_summary()
        if dist:
            rows = [r for r in rows if r.get('district') == dist]

        self.tbl_out_cust_ids = []
        self.tbl_out.setRowCount(len(rows))
        for r, row in enumerate(rows):
            self.tbl_out_cust_ids.append(row['customer_id'])
            outstanding = row['outstanding']
            color = "#ef4444" if outstanding > 0 else "#10b981"
            vals = [
                row['customer_name'], row.get('district', ''),
                fmt_currency(row.get('initial_ar', 0)),
                fmt_currency(row.get('sales_credit', 0)),
                fmt_currency(row['total_collected']),
                fmt_currency(outstanding),
            ]
            for c, val in enumerate(vals):
                item = QTableWidgetItem(str(val))
                item.setTextAlignment(Qt.AlignCenter)
                if c == 5:
                    item.setForeground(QColor(color))
                elif c == 2 and row.get('initial_ar', 0) != 0:
                    item.setForeground(QColor("#ea580c"))
                self.tbl_out.setItem(r, c, item)
        self.tbl_credit_det.setRowCount(0)

    def _export_ar_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
        except ImportError:
            QMessageBox.warning(self, "경고", "openpyxl 라이브러리가 설치되어 있지 않습니다.\npip install openpyxl 명령어로 설치해주세요.")
            return

        path, _ = QFileDialog.getSaveFileName(self, "엑셀 다운로드", "미수금현황.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "미수금현황"
        
        headers = ["거래처", "지역", "초기(이월) 미수", "판매 미수액", "총 수금액", "잔여 미수"]
        ws.append(headers)
        
        # 헤더 스타일
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        rows = self.ar_logic.get_outstanding_summary()
        dist = self.combo_dist_filter.currentData()
        if dist:
            rows = [r for r in rows if r.get('district') == dist]

        data_row = 2
        for row in rows:
            if row['outstanding'] <= 0:
                continue
            ws.append([
                row['customer_name'],
                row.get('district', ''),
                row.get('initial_ar', 0),
                row.get('sales_credit', 0),
                row['total_collected'],
                row['outstanding']
            ])
            # 데이터 셀 테두리 & 정렬
            for col in range(1, len(headers)+1):
                cell = ws.cell(row=data_row, column=col)
                cell.border = thin_border
                if col >= 3:  # 금액 열들
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            # 잔여 미수 셀 강조
            amt_cell = ws.cell(row=data_row, column=6)
            amt_cell.font = Font(bold=True, color="FF0000")
            data_row += 1

        # 셀 너비 자동 조정
        for col_cells in ws.columns:
            max_length = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                try:
                    cell_len = len(str(cell.value or ''))
                    if cell_len > max_length:
                        max_length = cell_len
                except:
                    pass
            ws.column_dimensions[col_letter].width = max(max_length + 4, 12)

        # 행 높이 조정
        for row_obj in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row_obj[0].row].height = 22

        try:
            wb.save(path)
            QMessageBox.information(self, "완료", f"엑셀 파일이 저장되었습니다.\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 저장 중 오류가 발생했습니다:\n{str(e)}")

    def _print_ar_list(self):
        """HTML 기반 미수금 현황 인쇄."""
        from PyQt5.QtPrintSupport import QPrinter, QPrintPreviewDialog
        from PyQt5.QtGui import QTextDocument
        from datetime import date as dt_date

        rows = self.ar_logic.get_outstanding_summary()
        dist = self.combo_dist_filter.currentData()
        if dist:
            rows = [r for r in rows if r.get('district') == dist]
        rows = [r for r in rows if r['outstanding'] > 0]

        today_str = dt_date.today().strftime('%Y-%m-%d')
        total_ar = sum(r['outstanding'] for r in rows)

        html = f"""
        <h2 style='text-align:center; margin-bottom:4px;'>미수금 현황 보고서</h2>
        <p style='text-align:center; color:#666; margin-top:0;'>출력일: {today_str} | 총 미수금: <b style='color:red;'>{fmt_currency(total_ar)}</b></p>
        <table border='1' cellspacing='0' cellpadding='5' width='100%' style='border-collapse:collapse; font-size:10pt;'>
        <tr style='background:#4472C4; color:white; font-weight:bold; text-align:center;'>
            <th>거래처</th><th>지역</th><th>초기(이월) 미수</th>
            <th>판매 미수액</th><th>총 수금액</th><th>잔여 미수</th>
        </tr>
        """
        for r in rows:
            html += f"""
            <tr style='text-align:center;'>
                <td>{r['customer_name']}</td>
                <td>{r.get('district','')}</td>
                <td style='text-align:right;'>{fmt_currency(r.get('initial_ar',0))}</td>
                <td style='text-align:right;'>{fmt_currency(r.get('sales_credit',0))}</td>
                <td style='text-align:right;'>{fmt_currency(r['total_collected'])}</td>
                <td style='text-align:right; color:red; font-weight:bold;'>{fmt_currency(r['outstanding'])}</td>
            </tr>
            """
        html += "</table>"

        doc = QTextDocument()
        doc.setHtml(html)

        def preview_paint(printer):
            doc.print_(printer)

        printer = QPrinter(QPrinter.HighResolution)
        printer.setPageOrientation(1)  # Landscape
        dialog = QPrintPreviewDialog(printer, self)
        dialog.setWindowTitle("미수금 현황 인쇄 미리보기")
        dialog.paintRequested.connect(preview_paint)
        dialog.exec_()

    def _on_select_customer(self):
        rows = self.tbl_out.selectedItems()
        if not rows:
            return
        r_idx = self.tbl_out.currentRow()
        if r_idx < 0 or r_idx >= len(self.tbl_out_cust_ids):
            return
        cid   = self.tbl_out_cust_ids[r_idx]
        sales = self.ar_logic.get_credit_sales_by_customer(cid)
        self.tbl_credit_det.setRowCount(len(sales))
        for r, row in enumerate(sales):
            vals = [row['sale_date'], row['type_name'],
                    row['spec_name'], fmt_number(row['quantity']),
                    fmt_currency(row['total_amount'])]
            for c, val in enumerate(vals):
                item = QTableWidgetItem(str(val))
                item.setTextAlignment(Qt.AlignCenter)
                self.tbl_credit_det.setItem(r, c, item)

    def _on_double_click_customer(self, row, col):
        """거래처 더블클릭 시 날짜별 그룹 상세 표시."""
        if row < 0 or row >= len(self.tbl_out_cust_ids):
            return
        cid = self.tbl_out_cust_ids[row]
        cust_name = self.tbl_out.item(row, 0).text() if self.tbl_out.item(row, 0) else ''
        date_rows = self.ar_logic.get_credit_by_date_grouped(cid)
        self.tbl_date_group.setRowCount(len(date_rows))
        for r, dr in enumerate(date_rows):
            vals = [
                dr['sale_date'],
                fmt_number(dr['item_count']) + '건',
                fmt_number(dr['total_qty']),
                fmt_currency(dr['total_amount'])
            ]
            for c, val in enumerate(vals):
                item = QTableWidgetItem(str(val))
                item.setTextAlignment(Qt.AlignCenter)
                if c == 3:  # 금액 강조
                    item.setForeground(QColor("#ef4444"))
                self.tbl_date_group.setItem(r, c, item)

    def _add_collection(self, cid=None):
        dlg = ARCollectionDialog(self, customer_id=cid)
        if dlg.exec_():
            self._load_outstanding()
            self._load_collections()
            self._load_summary()

    def _add_collection_selected(self):
        rows = self.tbl_out.selectedItems()
        if not rows:
            QMessageBox.information(self, "알림", "수금할 거래처를 선택해주세요.")
            return
        cid = self.tbl_out_cust_ids[self.tbl_out.currentRow()]
        self._add_collection(cid)

    def _adjust_initial_ar(self):
        rows = self.tbl_out.selectedItems()
        cid = None
        if rows:
            r_idx = self.tbl_out.currentRow()
            if 0 <= r_idx < len(self.tbl_out_cust_ids):
                cid = self.tbl_out_cust_ids[r_idx]
        dlg = InitialARDialog(self, customer_id=cid)
        if dlg.exec_():
            self._load_outstanding()
            self._load_summary()

    # ── 수금 내역 탭 ─────────────────────────────────────────
    def _collection_tab(self) -> QWidget:
        w = QWidget()
        v = QVBoxLayout(w)
        v.setContentsMargins(16, 14, 16, 14)
        v.setSpacing(12)

        fh = QHBoxLayout(); fh.setSpacing(8)

        def fl(t):
            l = QLabel(t); l.setObjectName("form_label"); return l

        fh.addWidget(fl("거래처"))
        self.combo_col_cust = QComboBox()
        self.combo_col_cust.setMinimumWidth(160)
        fh.addWidget(self.combo_col_cust)

        fh.addWidget(fl("시작"))
        self.date_col_s = NoScrollDateEdit(QDate.currentDate().addDays(-30))
        self.date_col_s.setCalendarPopup(True)
        self.date_col_s.setDisplayFormat("yyyy-MM-dd")
        self.date_col_s.setFixedWidth(130)
        fh.addWidget(self.date_col_s)

        fh.addWidget(fl("종료"))
        self.date_col_e = NoScrollDateEdit(QDate.currentDate())
        self.date_col_e.setCalendarPopup(True)
        self.date_col_e.setDisplayFormat("yyyy-MM-dd")
        self.date_col_e.setFixedWidth(130)
        fh.addWidget(self.date_col_e)

        btn_q = QPushButton("조회")
        btn_q.setObjectName("btn_ghost")
        btn_q.setFixedWidth(80)
        btn_q.clicked.connect(self._load_collections)
        fh.addWidget(btn_q)
        fh.addStretch()

        btn_add = QPushButton("수금 등록")
        btn_add.setObjectName("btn_success")
        btn_add.clicked.connect(lambda: self._add_collection(None))
        fh.addWidget(btn_add)
        v.addLayout(fh)

        c = SectionCard("수금 내역")
        self.tbl_col = self._tbl(
            ['날짜', '거래처', '지역', '수금액', '방법', '메모']
        )
        self._col_ids = []
        c.add_widget(self.tbl_col)

        dh = QHBoxLayout(); dh.addStretch()
        btn_del = QPushButton("선택 삭제")
        btn_del.setObjectName("btn_danger")
        btn_del.clicked.connect(self._delete_collection)
        dh.addWidget(btn_del)
        c.add_layout(dh)
        v.addWidget(c)

        self._load_col_combo()
        self._load_collections()
        return w

    def _load_col_combo(self):
        self.combo_col_cust.clear()
        self.combo_col_cust.addItem("전체", None)
        for c in self.customer_logic.get_all():
            self.combo_col_cust.addItem(c['name'], c['id'])

    def _load_collections(self):
        cid   = self.combo_col_cust.currentData() if hasattr(self, 'combo_col_cust') else None
        start = self.date_col_s.date().toString("yyyy-MM-dd") if hasattr(self, 'date_col_s') else None
        end   = self.date_col_e.date().toString("yyyy-MM-dd") if hasattr(self, 'date_col_e') else None
        rows  = self.ar_logic.get_collections(cid, start, end)
        METHOD_COLOR = {'현금':'#10b981','카드':'#6366f1','계좌이체':'#06b6d4'}
        self._col_ids = []
        self.tbl_col.setRowCount(len(rows))
        for r, row in enumerate(rows):
            self._col_ids.append(row['id'])
            vals = [row['collection_date'], row['customer_name'],
                    row.get('district',''), fmt_currency(row['amount']),
                    row['payment_method'], row.get('memo','')]
            for c, val in enumerate(vals):
                item = QTableWidgetItem(str(val))
                item.setTextAlignment(Qt.AlignCenter)
                if c == 4:
                    item.setForeground(QColor(METHOD_COLOR.get(val,'#e2e8f0')))
                if c == 3:
                    item.setForeground(QColor('#10b981'))
                self.tbl_col.setItem(r, c, item)

    def _delete_collection(self):
        sel = sorted(set(i.row() for i in self.tbl_col.selectedItems()))
        if not sel:
            QMessageBox.information(self, "알림", "삭제할 행을 선택해주세요.")
            return
        if QMessageBox.question(self, "삭제 확인",
            f"{len(sel)}건을 삭제하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return
        for r in sorted(sel, reverse=True):
            self.ar_logic.delete_collection(self._col_ids[r])
        self._load_collections()
        self._load_summary()
        self._load_outstanding()

    # ── 요약 ─────────────────────────────────────────────────
    def _load_summary(self):
        total    = self.ar_logic.get_outstanding_total()
        summary  = self.ar_logic.get_outstanding_summary()
        cust_cnt = len([r for r in summary if r['outstanding'] > 0])
        cur      = dt.today()
        monthly  = self.ar_logic.get_monthly_collection_total(cur.year, cur.month)
        self.sc_total.set_value(fmt_currency(total))
        self.sc_cust_cnt.set_value(f"{cust_cnt} 곳")
        self.sc_monthly.set_value(fmt_currency(monthly))

    def _tbl(self, headers) -> QTableWidget:
        tbl = QTableWidget(0, len(headers))
        tbl.setHorizontalHeaderLabels(headers)
        tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        tbl.setSelectionBehavior(QTableWidget.SelectRows)
        tbl.setAlternatingRowColors(True)
        tbl.verticalHeader().setVisible(False)
        tbl.setShowGrid(False)
        return tbl

    def refresh(self):
        self._load_col_combo()
        self._load_outstanding()
        self._load_collections()
        self._load_summary()
