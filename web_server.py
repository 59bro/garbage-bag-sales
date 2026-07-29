# ============================================================
# web_server.py  —  종량제 봉투 판매 관리 시스템 모바일 웹 API 서버
# ============================================================

import os
import sys
import io
import traceback
from datetime import datetime

# 프로젝트 루트 경로 추가
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from fastapi import FastAPI, HTTPException, Query, Body
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

# 로직 모듈 임포트 (기존 로직 100% 재사용)
from logic.customer_logic import CustomerLogic
from logic.product_logic import ProductLogic
from logic.sales_logic import SalesLogic
from logic.stock_logic import StockLogic
from logic.ar_logic import ARLogic
from logic.report_logic import ReportLogic

app = FastAPI(
    title="종량제 봉투 판매 관리 모바일 API",
    description="스마트폰/태블릿 반응형 웹 및 PC 프로그램 데이터 실시간 동기화 서버",
    version="1.1.0"
)

# CORS 설정 (모든 도메인 허용)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 정적 파일 서빙 (CSS, JS)
static_dir = os.path.join(os.path.dirname(__file__), "static")
templates_dir = os.path.join(os.path.dirname(__file__), "templates")
os.makedirs(static_dir, exist_ok=True)
os.makedirs(templates_dir, exist_ok=True)

app.mount("/static", StaticFiles(directory=static_dir), name="static")

# 로직 객체 생성
customer_logic = CustomerLogic()
product_logic  = ProductLogic()
sales_logic     = SalesLogic()
stock_logic     = StockLogic()
ar_logic        = ARLogic()
report_logic    = ReportLogic()


# ── 데이터 모델 정의 (Pydantic) ─────────────────────────────
class SaleCreateSchema(BaseModel):
    sale_date: str
    customer_id: int
    spec_id: int
    quantity: int
    unit_price: int
    payment_method: str = "미수"
    cash_amount: int = 0
    card_amount: int = 0
    memo: str = ""

class CustomerCreateSchema(BaseModel):
    name: str
    district: str = ""
    phone: str = ""
    address: str = ""
    memo: str = ""
    customer_type: str = "출고처"
    initial_ar: int = 0

class InboundCreateSchema(BaseModel):
    transaction_date: str
    spec_id: int
    quantity: int
    memo: str = ""

class ARCollectionSchema(BaseModel):
    collection_date: str
    customer_id: int
    amount: int
    payment_method: str = "현금"
    memo: str = ""


# ── 페이지 루트 엔드포인트 ──────────────────────────────────
@app.get("/", response_class=HTMLResponse)
async def get_index():
    index_file = os.path.join(templates_dir, "index.html")
    if os.path.exists(index_file):
        return FileResponse(index_file)
    return HTMLResponse("<h1>웹 프론트엔드 templates/index.html 파일을 찾을 수 없습니다.</h1>", status_code=404)


# ── 대시보드 API ──────────────────────────────────────────────
@app.get("/api/dashboard")
async def get_dashboard_summary():
    try:
        today_str = datetime.now().strftime("%Y-%m-%d")
        month_prefix = datetime.now().strftime("%Y-%m")

        today_sales = sales_logic.get_sales_by_date(today_str)
        today_total = sum(r['total_amount'] for r in today_sales)
        today_qty   = sum(r['quantity'] for r in today_sales)

        total_ar = ar_logic.get_outstanding_total()

        month_sales = sales_logic.get_sales_by_period(f"{month_prefix}-01", today_str)
        month_total = sum(r['total_amount'] for r in month_sales)

        return {
            "status": "success",
            "today_date": today_str,
            "today_total_amount": today_total,
            "today_total_quantity": today_qty,
            "today_sales_count": len(today_sales),
            "total_ar_balance": total_ar,
            "month_total_amount": month_total
        }
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ── 기초 마스터 정보 API (거래처, 품목) ─────────────────────────
@app.get("/api/products")
async def get_products():
    try:
        specs = product_logic.get_all_specs(include_inactive=False)
        types = product_logic.get_types()
        return {
            "status": "success",
            "types": types,
            "specs": specs
        }
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/customers")
async def get_customers(district: str = None, customer_type: str = None):
    try:
        custs = customer_logic.get_all(include_inactive=True, district=district, customer_type=customer_type)
        return {"status": "success", "customers": custs}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/customers")
async def create_customer(data: CustomerCreateSchema):
    try:
        cid = customer_logic.add(
            name=data.name,
            district=data.district,
            phone=data.phone,
            address=data.address,
            memo=data.memo,
            customer_type=data.customer_type,
            initial_ar=data.initial_ar
        )
        return {"status": "success", "customer_id": cid, "message": "거래처가 등록되었습니다."}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/customers/{customer_id}")
async def delete_customer(customer_id: int):
    try:
        ref = customer_logic.has_references(customer_id)
        if ref['has_ref']:
            return {
                "status": "warning",
                "can_delete": False,
                "ref_info": ref,
                "message": f"기존 거래/수금 내역({ref['total_cnt']}건)이 있어 삭제할 수 없습니다. 비활성화를 이용하세요."
            }
        customer_logic.delete(customer_id)
        return {"status": "success", "can_delete": True, "message": "거래처가 완전 삭제되었습니다."}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/customers/{customer_id}/deactivate")
async def deactivate_customer(customer_id: int):
    try:
        customer_logic.deactivate(customer_id)
        return {"status": "success", "message": "거래처가 비활성화되었습니다."}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ── 판매 관리 API ──────────────────────────────────────────────
@app.get("/api/sales")
async def get_sales(
    date: str = None,
    start_date: str = None,
    end_date: str = None,
    customer_id: int = None
):
    try:
        if date:
            rows = sales_logic.get_sales_by_date(date)
        elif start_date and end_date:
            rows = sales_logic.get_sales_by_period(start_date, end_date, customer_id=customer_id)
        else:
            today_str = datetime.now().strftime("%Y-%m-%d")
            rows = sales_logic.get_sales_by_date(today_str)
        return {"status": "success", "sales": rows}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/sales")
async def create_sale(data: SaleCreateSchema):
    try:
        sid = sales_logic.add_sale(
            sale_date=data.sale_date,
            customer_id=data.customer_id,
            spec_id=data.spec_id,
            quantity=data.quantity,
            unit_price=data.unit_price,
            payment_method=data.payment_method,
            memo=data.memo,
            cash_amount=data.cash_amount,
            card_amount=data.card_amount
        )
        return {"status": "success", "sale_id": sid, "message": "판매 내역이 등록되었습니다."}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/sales/{sale_id}")
async def delete_sale(sale_id: int):
    try:
        sales_logic.delete_sale(sale_id)
        return {"status": "success", "message": "판매 내역이 삭제되었습니다."}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ── 재고 관리 API ──────────────────────────────────────────────
@app.get("/api/stock")
async def get_stock_status():
    try:
        stocks = stock_logic.get_all_current_stocks()
        return {"status": "success", "stock": stocks}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/stock/inbound")
async def add_inbound_stock(data: InboundCreateSchema):
    try:
        iid = stock_logic.add_inbound(
            spec_id=data.spec_id,
            quantity=data.quantity,
            date=data.transaction_date,
            memo=data.memo
        )
        return {"status": "success", "inbound_id": iid, "message": "입고 내역이 등록되었습니다."}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ── 미수 관리 (AR) API ─────────────────────────────────────────
@app.get("/api/ar")
async def get_ar_balances(search: str = None):
    try:
        summary = ar_logic.get_outstanding_summary()
        balances = [
            {
                "id": row["customer_id"],
                "name": row["customer_name"],
                "district": row["district"],
                "ar_balance": row["outstanding"]
            }
            for row in summary
            if not search or search.lower() in row["customer_name"].lower()
        ]
        return {"status": "success", "ar_balances": balances}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/ar/collection")
async def add_ar_collection(data: ARCollectionSchema):
    try:
        cid = ar_logic.add_collection(
            collection_date=data.collection_date,
            customer_id=data.customer_id,
            amount=data.amount,
            payment_method=data.payment_method,
            memo=data.memo
        )
        return {"status": "success", "collection_id": cid, "message": "수금 내역이 등록되었습니다."}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/ar/export")
async def export_ar_to_excel():
    """미수금 현황을 엑셀 파일로 다운로드."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        summary = ar_logic.get_outstanding_summary()
        wb = Workbook()
        ws = wb.active
        ws.title = "미수금현황"

        # 스타일 정의
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        center = Alignment(horizontal="center", vertical="center")
        right = Alignment(horizontal="right", vertical="center")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        money_fmt = '#,##0'

        # 타이틀
        today_str = datetime.now().strftime('%Y-%m-%d %H:%M')
        ws.merge_cells('A1:F1')
        ws['A1'] = f"미수금 현황 - 출력일: {today_str}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center

        # 헤더
        headers = ['거래처명', '지역', '초기미수', '판매미수', '수금액', '미수잔액']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        # 데이터
        total_outstanding = 0
        for i, row in enumerate(summary, 4):
            ws.cell(row=i, column=1, value=row['customer_name']).border = border
            ws.cell(row=i, column=2, value=row['district']).border = border
            c3 = ws.cell(row=i, column=3, value=row['initial_ar'])
            c3.number_format = money_fmt; c3.alignment = right; c3.border = border
            c4 = ws.cell(row=i, column=4, value=row['sales_credit'])
            c4.number_format = money_fmt; c4.alignment = right; c4.border = border
            c5 = ws.cell(row=i, column=5, value=row['total_collected'])
            c5.number_format = money_fmt; c5.alignment = right; c5.border = border
            c6 = ws.cell(row=i, column=6, value=row['outstanding'])
            c6.number_format = money_fmt; c6.alignment = right; c6.border = border
            c6.font = Font(bold=True, color="E74C3C" if row['outstanding'] > 0 else "27AE60")
            total_outstanding += row['outstanding']

        # 합계행
        sum_row = len(summary) + 4
        ws.cell(row=sum_row, column=1, value='합계').font = Font(bold=True, size=12)
        c_total = ws.cell(row=sum_row, column=6, value=total_outstanding)
        c_total.number_format = money_fmt
        c_total.font = Font(bold=True, size=12, color="E74C3C")
        c_total.alignment = right

        # 열 너비
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        for c in 'CDEF':
            ws.column_dimensions[c].width = 16

        # 스트리밍 응답
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"미수금현황_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
        )
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    if hasattr(sys.stdout, 'reconfigure'):
        try:
            sys.stdout.reconfigure(encoding='utf-8')
        except Exception:
            pass
    print("=====================================================================")
    print("  종량제 봉투 판매 관리 모바일 웹 서버 실행 시작")
    print("  PC 웹 접속 주소: http://127.0.0.1:8000")
    print("  모바일(동일 Wi-Fi) 접속 주소: http://0.0.0.0:8000")
    print("=====================================================================")
    uvicorn.run("web_server:app", host="0.0.0.0", port=8000, reload=True)
