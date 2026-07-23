# ============================================================
# utils/export_utils.py
# 엑셀 내보내기 및 특정 양식(일일판매일지.xls) 자동 입력 유틸
# ============================================================

import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import xlrd
    from xlutils.copy import copy
    import xlwt
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False


def export_to_excel(headers: list, rows: list, title: str = "판매현황", filepath: str = None) -> str | None:
    """
    데이터를 엑셀로 내보내기 (일반 목록형 openpyxl).
    filepath를 전달하면 해당 경로에 저장합니다.
    반환값: 저장된 파일 경로 (실패시 None)
    """
    if not EXCEL_AVAILABLE:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:30]

    # 스타일 정의
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 타이틀 행
    ws.merge_cells(f"A1:{chr(64 + len(headers))}1")
    ws["A1"] = f"{title} - 출력일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 25

    # 헤더 행
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    ws.row_dimensions[2].height = 20

    # 데이터 행
    for row_idx, row_data in enumerate(rows, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_align
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    # 열 너비 자동 조정
    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ''))
            for r in range(1, len(rows) + 3)
        )
        ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 4, 30)

    # 저장 경로
    if not filepath:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        if not os.path.exists(desktop):
            desktop = os.path.join(os.path.expanduser("~"), "Documents")
        filename = f"{title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(desktop, filename)

    wb.save(filepath)
    return filepath


def map_spec_to_col(type_name: str, spec_name: str, product_code: str = '') -> int | None:
    """
    (품종명, 규격명, 상품코드)을 일일판매일지 양식의 열 인덱스(0-based, c=3~25)로 매핑.
    상품코드에 'D', 'E' 등 열 알파벳이 입력되어 있으면 해당 열로 우선 매핑합니다.
    """
    code = str(product_code).strip().upper().replace('COL:', '').replace('열:', '')
    if len(code) == 1 and 'D' <= code <= 'Z':
        return ord(code) - ord('A')
    if code.isdigit() and 3 <= int(code) <= 25:
        return int(code)

    t = type_name.replace('용 봉투', '').replace('봉투', '').strip()
    s = str(spec_name).replace('L', '').replace('ℓ', '').strip()

    if '생활' in t:
        if s == '5': return 3
        if s == '10': return 4
        if s == '20': return 5
        if s == '30': return 6
        if s == '50': return 7
        if s == '75': return 8
    elif '재사용' in t:
        if s == '20': return 9
        if s == '10': return 10
    elif '반장' in t:
        if s == '10': return 11
    elif '비매' in t:
        if s == '3': return 12
    elif '음식물' in t:
        if s == '1': return 13
        if s == '2': return 14
        if s == '3': return 15
        if s == '5': return 16
        if s == '10': return 17
    elif '업소' in t or ('필증' in t and '가정' not in t):
        if s == '5': return 18
        if s == '10': return 19
        if s == '20': return 20
        if s == '60': return 21
        if s == '120': return 22
    elif '가정' in t:
        if s == '120': return 23
    elif '마대' in t:
        if s == '10': return 24
        if s == '20': return 25
    return None


def _write_preserve_style(sheet, r: int, c: int, val):
    """
    xlutils 복사본 시트에서 기존 셀의 스타일(xf_idx)을 유지하면서 값을 작성.
    """
    row_obj = sheet._Worksheet__rows.get(r)
    old_xf = None
    if row_obj and c in row_obj._Row__cells:
        old_xf = row_obj._Row__cells[c].xf_idx

    sheet.write(r, c, val)

    if old_xf is not None:
        sheet._Worksheet__rows[r]._Row__cells[c].xf_idx = old_xf


def _write_cloned_10pt(sheet, rb_book, sh_read, r: int, c: int, val, font_size: int = 10):
    """
    템플릿 엑셀의 기존 셀(r, c)이 가진 모든 서식(테두리, 색상, 정렬, 패턴 등)을
    100% 그대로 복제한 후, 글자 크기만 10pt로 변경하여 값을 작성합니다.
    이를 통해 다른 셀들과 테두리/서식 이질감이 전혀 발생하지 않습니다.
    """
    if not val and val != 0 and val != 0.0:
        _write_preserve_style(sheet, r, c, val)
        return

    try:
        old_idx = sh_read.cell_xf_index(r, c)
        old_xf = rb_book.xf_list[old_idx]
        old_font = rb_book.font_list[old_xf.font_index]

        style = xlwt.XFStyle()

        # 1. 폰트 복제 (크기만 10pt -> 200)
        font = xlwt.Font()
        font.name = old_font.name
        font.height = font_size * 20  # 10pt = 200
        font.bold = old_font.bold
        font.italic = old_font.italic
        font.underline = old_font.underline_type
        font.colour_index = old_font.colour_index
        style.font = font

        # 2. 테두리 스타일 및 색상 완벽 복제
        style.borders.left = old_xf.border.left_line_style
        style.borders.right = old_xf.border.right_line_style
        style.borders.top = old_xf.border.top_line_style
        style.borders.bottom = old_xf.border.bottom_line_style
        style.borders.left_colour = old_xf.border.left_colour_index
        style.borders.right_colour = old_xf.border.right_colour_index
        style.borders.top_colour = old_xf.border.top_colour_index
        style.borders.bottom_colour = old_xf.border.bottom_colour_index

        # 3. 배경색(패턴) 완벽 복제
        style.pattern.pattern = old_xf.background.fill_pattern
        style.pattern.pattern_fore_colour = old_xf.background.pattern_colour_index
        style.pattern.pattern_back_colour = old_xf.background.background_colour_index

        # 4. 정렬 완벽 복제 (금액 열인 26, 27, 28열은 회계식 우측 정렬 적용)
        if c in (26, 27, 28):
            style.alignment.horz = xlwt.Alignment.HORZ_RIGHT  # 3 (우측 정렬)
            style.alignment.vert = xlwt.Alignment.VERT_CENTER # 1 (수직 중앙)
        else:
            style.alignment.horz = old_xf.alignment.hor_align
            style.alignment.vert = old_xf.alignment.vert_align
        style.alignment.wrap = old_xf.alignment.text_wrapped

        # 5. 천단위 콤마 회계식 숫자 형식 적용
        if old_xf.format_key in rb_book.format_map:
            style.num_format_str = rb_book.format_map[old_xf.format_key].format_str
        elif isinstance(val, (int, float)):
            style.num_format_str = '#,##0'

        sheet.write(r, c, val, style)
    except Exception:
        _write_preserve_style(sheet, r, c, val)


def export_daily_sales_template(date_str: str, sales_data: list,
                                template_path: str = None, output_path: str = None,
                                collections_data: list = None) -> str:
    """
    특정 일일판매일지 양식(F:\\일일판매일지\\엑셀연동일일판매일지.xls)에
    DB의 날짜별 판매 및 미수 수금(결제) 데이터를 완벽하게 매칭하여 입력 후 새 파일로 저장.
    
    date_str: "2026-07-14" 형식
    sales_data: logic.get_sales_by_date(date_str) 결과 행 리스트
    collections_data: logic.ar_logic.get_collections(start_date=date_str, end_date=date_str) 결과 행 리스트
    """
    if not XLRD_AVAILABLE:
        raise RuntimeError("xlrd 및 xlutils 라이브러리가 설치되지 않았습니다.")

    if not template_path:
        from utils.db_config_manager import get_root_dir
        from database.db_manager import DBManager
        candidates = [
            r'F:\일일판매일지\엑셀연동일일판매일지.xls',
            os.path.join(os.path.dirname(DBManager().db_path), '엑셀연동일일판매일지.xls'),
            os.path.join(get_root_dir(), '엑셀연동일일판매일지.xls'),
            os.path.join(get_root_dir(), 'data', '엑셀연동일일판매일지.xls'),
            r'G:\내 드라이브\종량제봉투_통합DB\엑셀연동일일판매일지.xls',
            r'D:\내 드라이브\종량제봉투_통합DB\엑셀연동일일판매일지.xls'
        ]
        for c in candidates:
            if os.path.exists(c):
                template_path = c
                break
        if not template_path:
            raise RuntimeError("기본 엑셀 양식 파일(엑셀연동일일판매일지.xls)을 찾을 수 없습니다. DB 폴더 또는 실행 파일 폴더에 양식을 넣어주세요.")

    try:
        rb = xlrd.open_workbook(template_path, formatting_info=True)
    except Exception as e:
        raise RuntimeError("선택한 엑셀 파일이 .xlsx 형식이거나 지원되지 않습니다. .xls 97-2003 양식을 사용해주세요.")

    wb_wt = copy(rb)
    wb = wb_wt
    s_wt = wb_wt.get_sheet(0)
    sh_read = rb.sheet_by_index(0)

    # 페이지 설정: 가로 방향 / 확대·축소 60%
    try:
        s_wt.portrait = False
        s_wt.print_scaling = 60
    except Exception:
        pass

    # 날짜 셀: Row 2 (3번째 줄, 0-based), Col 2 = "2026- 7 -", Col 3 = 14
    # Row 0(타이틀: 일일판매일지), Row 1(빈 행), Row 2(날짜 필드)
    try:
        dt_obj = datetime.strptime(date_str, "%Y-%m-%d")
        year_month_str = f"{dt_obj.year}- {dt_obj.month:2d} -"
        _write_preserve_style(s_wt, 2, 2, year_month_str)
        _write_preserve_style(s_wt, 2, 3, dt_obj.day)
    except Exception:
        pass

    # 2. 거래처별로 판매 및 수금 데이터 그룹화
    customer_groups = {}
    for row in sales_data:
        cname = row['customer_name']
        if cname not in customer_groups:
            customer_groups[cname] = []
        customer_groups[cname].append(row)

    collection_groups = {}
    if collections_data:
        for col in collections_data:
            cname = col['customer_name']
            if cname not in collection_groups:
                collection_groups[cname] = []
            collection_groups[cname].append(col)

    all_customers = list(customer_groups.keys())
    for cname in collection_groups.keys():
        if cname not in customer_groups and cname not in all_customers:
            all_customers.append(cname)

    current_row = 6
    max_data_row = 46  # 47행(Row 48)이 합계행이므로 46행까지 작성 가능

    col_sums = {c: 0.0 for c in range(3, 29)}

    # 3. 데이터 행 작성
    for cname in all_customers:
        if current_row > max_data_row:
            break

        _write_preserve_style(s_wt, current_row, 2, cname)

        row_counts = {c: 0 for c in range(3, 26)}
        total_amt = 0.0
        cash_amt = 0.0
        card_amt = 0.0

        # 판매 내역 처리
        for item in customer_groups.get(cname, []):
            col_idx = map_spec_to_col(item['type_name'], item['spec_name'], item.get('product_code', ''))
            if col_idx is not None and col_idx in row_counts:
                row_counts[col_idx] += item['quantity']

            amt = float(item['total_amount'])
            total_amt += amt
            if item['payment_method'] == '현금':
                cash_amt += amt
            elif item['payment_method'] == '카드':
                card_amt += amt

        # 미수 수금(결제) 내역 처리: 판매액(col 26) 외에 결제한 수금액을 현금/카드 열에 합산!
        for col in collection_groups.get(cname, []):
            amt = float(col['amount'])
            method = col.get('payment_method', '현금')
            if method in ('현금', '계좌이체'):
                cash_amt += amt
            elif method == '카드':
                card_amt += amt

        # 수량 열 작성
        for c in range(3, 26):
            qty = row_counts[c]
            if qty > 0:
                _write_preserve_style(s_wt, current_row, c, qty)
                col_sums[c] += qty
            else:
                _write_preserve_style(s_wt, current_row, c, '')

        # 판매액 / 현금 / 카드 열 작성 (템플릿 서식/테두리 100% 복제 + 폰트 크기 10pt)
        _write_cloned_10pt(s_wt, rb, sh_read, current_row, 26, total_amt if total_amt > 0 else '', font_size=10)
        _write_cloned_10pt(s_wt, rb, sh_read, current_row, 27, cash_amt if cash_amt > 0 else '', font_size=10)
        _write_cloned_10pt(s_wt, rb, sh_read, current_row, 28, card_amt if card_amt > 0 else '', font_size=10)

        if total_amt > 0:
            col_sums[26] += total_amt
        if cash_amt > 0:
            col_sums[27] += cash_amt
        if card_amt > 0:
            col_sums[28] += card_amt

        current_row += 1

    # 4. 남은 빈 행(입력된 거래처 다음 행부터 46행까지) 깔끔하게 초기화 (기존 템플릿 테두리 그대로 유지)
    for r in range(current_row, max_data_row + 1):
        for c in range(2, 29):
            _write_preserve_style(s_wt, r, c, '')

    # 5. 합계 행(Row 48, 인덱스 47)에 각 열의 총 합계 작성
    sum_row = 47
    for c in range(3, 29):
        val = col_sums[c]
        if c in (26, 27, 28):
            _write_cloned_10pt(s_wt, rb, sh_read, sum_row, c, val if val > 0 else 0.0, font_size=10)
        else:
            if val > 0:
                _write_preserve_style(s_wt, sum_row, c, val)
            else:
                _write_preserve_style(s_wt, sum_row, c, 0.0)

    # 6. 새 파일로 저장
    if not output_path:
        out_dir = r'F:\일일판매일지' if os.path.exists(r'F:\일일판매일지') else os.path.join(os.path.expanduser("~"), "Desktop")
        dt_clean = date_str.replace('-', '')
        output_path = os.path.join(out_dir, f"일일판매일지_{dt_clean}.xls")

    try:
        wb_wt.save(output_path)
    except (PermissionError, OSError) as e:
        if "Permission denied" in str(e) or "13" in str(e) or isinstance(e, PermissionError):
            raise RuntimeError(
                f"엑셀 파일('{os.path.basename(output_path)}')이 현재 다른 프로그램(또는 엑셀)에서 열려있어 덮어쓸 수 없습니다.\n\n"
                "해당 엑셀 파일을 닫으신 후 다시 시도해주시거나, 다른 파일 이름으로 저장해주세요."
            )
        raise e
    return output_path


def print_excel_file(filepath: str) -> bool:
    """
    생성된 엑셀 파일을 윈도우 기본 프린터로 즉시 인쇄 (가로 방향 2, 축소 60% 자동 적용).
    """
    try:
        import win32com.client
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(filepath)
        try:
            ws = wb.Sheets(1)
            ws.PageSetup.Orientation = 2  # xlLandscape
            ws.PageSetup.Zoom = 60
        except Exception:
            pass
        wb.PrintOut()
        wb.Close(SaveChanges=False)
        excel.Quit()
        return True
    except Exception:
        try:
            os.startfile(filepath, "print")
            return True
        except Exception as e:
            raise RuntimeError(f"프린터 출력 실패: {e}")
